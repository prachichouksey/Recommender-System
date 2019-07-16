# -*- coding: utf-8 -*-
"""
Created on Tue Jun 25 15:03:38 2019

@author: Ruchika
"""

# coding: utf-8

from bs4 import BeautifulSoup
import requests
import re
import pandas as pd


class Constants:
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'
    request_header = {'user-agent': user_agent}


class Patent:
    def __init__(self, title: str, url: str):
        self.title = title
        self.url = url
        self.fetched_details = False
        self.patent_num = None
        self.patent_date = None
        self.cpc=None
        self.file_date = None
        self.abstract = None
#        self.description = None
        self.inventors = None
        self.applicant_num = None
        self.applicant_name = None
        self.applicant_city = None
        self.applicant_state = None
        self.applicant_country = None
        self.assignee_name = None
        self.assignee_loc = None
        self.family_id = None
#        self.claims = None

    def fetch_details(self):
        self.fetched_details = True
        r = requests.get(self.url, headers=Constants.request_header).text
        s = BeautifulSoup(r, 'html.parser')
        try:
            self.patent_num = s.find(string='United States Patent ').find_next().text.replace('\n', '').strip().replace(',','')
        except:
            pass

        try:
            self.patent_date = s.find_all(align='right', width='50%')[-1].text.replace('\n', '').strip()
        except:
            pass

        try:
            abstract = s.find(string='Abstract').find_next().text.replace('\n', '').strip()
            self.abstract = re.sub(' +',' ', abstract)
        except:
            pass

        try:
            inventors = s.find(string='Inventors:').find_next().text.replace('\n', '').split('),')
            inventors = [t.split(';') for t in inventors]
            inventors = [[i.split('(') for i in j] for j in inventors]
            self.inventors = []
            for person in inventors:
                lname = person[0][0].strip()
                fname = person[1][0].strip()
                loc = person[1][1].strip().replace(')', '')
                d = [fname, lname, loc]
                self.inventors.append(d)
        except:
            pass
        
        

        try:
            self.applicant_name = s.find(string=re.compile('Applicant:')).find_next().find('td').text.replace('\n', '').strip()
        except:
            pass

        try:
            rem_applicant_data = s.find(string=re.compile('Applicant:')).find_next().find('td').find_next_siblings()
            try:
                self.applicant_city = rem_applicant_data[0].text.replace('\n', '').strip()
            except:
                pass
            try:
                self.applicant_state = rem_applicant_data[1].text.replace('\n', '').strip()
            except:
                pass
            try:
                self.applicant_country = rem_applicant_data[2].text.replace('\n', '').strip()
            except:
                pass
        except:
            pass

        try:
            assignee_raw = s.find(string=re.compile('Assignee:')).find_next().text.replace('\n', '')
            assignee_data = assignee_raw.split('(')
            try:
                self.assignee_name = assignee_data[0].strip()
            except:
                pass
            try:
                self.assignee_loc = assignee_data[1].strip().replace(')', '')
            except:
                pass
        except:
            pass

        try:
            self.family_id = s.find(string=re.compile('Family ID:')).find_next().text.replace('\n', '').strip()
        except:
            pass

        try:
            self.applicant_num = s.find(string=re.compile('Appl. No.:')).find_next().text.replace('\n', '').strip()
        except:
            pass

        try:
            self.file_date = s.find(string=re.compile('Filed:')).find_next().text.strip()
        except:
            pass
        
        try:
            cpc=s.find(string=re.compile('Current CPC Class:')).find_next().text.replace('\n', '').split(';,')
            cpc = [t.split(';') for t in cpc]
            mcpc = []
            for cpc_class in cpc:
                for mcpc_class in cpc_class:
                    mcpc.append(mcpc_class[0:5].strip())
            self.cpc = list( dict.fromkeys(mcpc))
#            print(self.cpc)
        except:
            pass
#        try:
#            claims = s.find(string=re.compile('Claims')).find_all_next(string=True)
#            claims = claims[:claims.index('Description')]
#            self.claims = [i.replace('\n', '').strip() for i in claims if i.replace('\n', '').strip() != '']
#        except:
#            pass

#        try:
#            description = s.find(string=re.compile('Description')).find_all_next(string=True)
#            self.description = [i.replace('\n', '').strip() for i in description if i.replace('\n', '').strip() not in ['', '* * * * *']]
#        except:
#            pass

    def as_dict(self) -> dict:
        """
        Return patent info as a dict
        :return: dict
        """
        if self.fetched_details:
            d = {
                'title': self.title,
                'patent_num': self.patent_num,
                'patent_date': self.patent_date,
                'file_date': self.file_date,
                'abstract': self.abstract,
                'cpc': self.cpc,
                'inventors': self.inventors,
                'applicant_name': self.applicant_name,
                'applicant_city': self.applicant_city,
                'applicant_state': self.applicant_state,
                'applicant_country': self.applicant_country,
                'assignee_name': self.assignee_name,
                'assignee_loc': self.assignee_loc,
                'family_id': self.family_id,
                'applicant_num': self.applicant_num,
#                'claims': self.claims,
#                'description': self.description,
                'url': self.url
            }
        else:
            d = {
                'title': self.title,
                'url': self.url
            }

        return d

    def __repr__(self):
        return str(self.as_dict())


class Search:
    def __init__(self,
                 string=None,
                 results_limit=1000,
                 get_patent_details=True,
                 pn=None,
                 isd=None,
                 ttl=None,
                 abst=None,
                 aclm=None,
                 spec=None,
                 ccl=None,
                 cpc=None,
                 cpcl=None,
                 icl=None,
                 apn=None,
                 apd=None,
                 apt=None,
                 govt=None,
                 fmid=None,
                 parn=None,
                 rlap=None,
                 rlfd=None,
                 prir=None,
                 prad=None,
                 pct=None,
                 ptad=None,
                 pt3d=None,
                 pppd=None,
                 reis=None,
                 rpaf=None,
                 afff=None,
                 afft=None,
                 in_=None,
                 ic=None,
                 is_=None,
                 icn=None,
                 aanm=None,
                 aaci=None,
                 aast=None,
                 aaco=None,
                 aaat=None,
                 lrep=None,
                 an=None,
                 ac=None,
                 as_=None,
                 acn=None,
                 exp=None,
                 exa=None,
                 ref=None,
                 fref=None,
                 oref=None,
                 cofc=None,
                 reex=None,
                 ptab=None,
                 sec=None,
                 ilrn=None,
                 ilrd=None,
                 ilpd=None,
                 ilfd=None):
        self.get_patent_details = get_patent_details
        args = {k: str(v).replace(' ', '-') for k, v in locals().items() if v and v is not self and v not in [get_patent_details, results_limit]}
        searchstring = ' AND '.join(['%s/%s' % (key, value) for (key, value) in args.items() if key not in ['results_limit']])
        searchstring = searchstring.replace('string/', '')
        searchstring = searchstring.replace(' ', '+')

        replace_dict = {'/': '%2F'}

        for k, v in replace_dict.items():
            searchstring = searchstring.replace(k, v)

        base_url = 'http://patft.uspto.gov/netacgi/nph-Parser?Sect1=PTO2&Sect2=HITOFF&u=%2Fnetahtml%2FPTO%2Fsearch-adv.htm&r=0&p=1&f=S&l=50&Query='

        url = base_url + searchstring + '&d=PTXT'
        print(url)
        r = requests.get(url, headers=Constants.request_header).text
        s = BeautifulSoup(r, 'html.parser')
        total_results = int(s.find(string=re.compile('out of')).find_next().text.strip())

        patents = self.get_patents_from_results_url(url, limit=results_limit)

        num_results_fetched = len(patents)

        list_num = 2

        base_url_nextpgs = 'http://patft.uspto.gov/netacgi/nph-Parser?Sect1=PTO2&Sect2=HITOFF&u=%2Fnetahtml%2FPTO%2Fsearch-adv.htm&r=0&f=S&l=50&d=PTXT'

        url_pre = base_url_nextpgs + '&OS=' + searchstring + '&RS=' + searchstring + '&Query=' + searchstring + '&TD=' + str(total_results) + '&Srch1=' + searchstring + '&NextList'
        url_post = '=Next+50+Hits'

        while (num_results_fetched < total_results) and (num_results_fetched < results_limit):
            this_url = url_pre + str(list_num) + url_post
            thispatents = self.get_patents_from_results_url(this_url)
            patents.extend(thispatents)

            num_results_fetched = len(patents)

            if num_results_fetched >= results_limit:
                patents = patents[:results_limit]

            list_num += 1

        self.patents = patents

    def get_patents_from_results_url(self, url: str, limit: int = None) -> list:
        r = requests.get(url, headers=Constants.request_header).text
        s = BeautifulSoup(r, 'html.parser')
        patents_raw = s.find_all('a', href=re.compile('netacgi'))
        patents_base_url = 'http://patft.uspto.gov'
        patents_raw_list = [[i.text.replace('\n', '').strip(), patents_base_url + i['href']] for i in patents_raw if
                            i.text.replace('\n', '').strip() != '']

        patents = []

        for patent_num_idx in range(0, len(patents_raw_list), 2):
            if limit and (patent_num_idx + 1) > limit:
                break
            patent_title = patents_raw_list[patent_num_idx + 1][0]
            patent_title = re.sub(' +', ' ', patent_title)
            patent_link = patents_raw_list[patent_num_idx][1]
            p = Patent(patent_title, patent_link)
            if self.get_patent_details:
                p.fetch_details()
            patents.append(p)

        return patents

    def as_dataframe(self) -> pd.DataFrame:
        if self.get_patent_details:
            return pd.DataFrame.from_records(self.as_list(), index='patent_num')
        else:
            return pd.DataFrame.from_records(self.as_list())

    def as_list(self) -> list:
        return [p.as_dict() for p in self.patents]

    def __repr__(self):
        return str(self.as_dataframe())
    
if __name__ == '__main__':
    import pandas
    from openpyxl import load_workbook
    s = Search('MANET').as_dataframe()
    writer = pandas.ExcelWriter('test4.xlsx', engine='openpyxl')
    writer.book = load_workbook('test4.xlsx')
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    print(writer.sheets['Sheet1'].max_row)
#    s.to_excel("test4.xlsx")
    s.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, header = False)
    writer.save()
    
    
    
    